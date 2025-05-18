import openpyxl
from openpyxl import Workbook, load_workbook
import os

# Corregido: apunta a la carpeta 'data' en la raíz del proyecto
DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../data'))

# Definición de cabeceras para cada planilla
PLANTILLAS = {
    'vehiculos.xlsx': [
        'id', 'tipo', 'marca', 'modelo', 'año', 'patente', 'estado', 'chofer_asignado',
        'fecha_ultimo_servicio', 'km_actual', 'permiso_circulacion_venc', 'revision_tecnica_venc',
        'soap_venc', 'decreto_60_venc', 'mantencion_venc', 'observaciones',
        'proveedor', 'fecha_alta', 'fecha_baja'
    ],
    'reservas.xlsx': [
        'id', 'fecha_solicitud', 'fecha_inicio', 'fecha_fin', 'vehiculo_id', 'solicitado_por',
        'cliente_obra', 'destino', 'chofer', 'estado', 'motivo', 'observaciones'
    ],
    'servicios.xlsx': [
        'id', 'fecha', 'vehiculo_id', 'chofer', 'cliente_obra', 'origen', 'destino',
        'km_inicio', 'km_fin', 'duracion', 'estado', 'observaciones'
    ],
    'pagos.xlsx': [
        'id', 'fecha_emision', 'fecha_vencimiento', 'cliente_proveedor', 'monto', 'concepto',
        'pagado', 'fecha_pago', 'forma_pago', 'observaciones'
    ],
    'personal.xlsx': [
        'id', 'nombre', 'rut', 'licencia', 'fecha_venc_licencia', 'contacto', 'estado', 'observaciones'
    ],
    'proveedores.xlsx': [
        'id', 'nombre', 'rut', 'contacto', 'tipo_servicio', 'estado', 'observaciones'
    ]
}

def crear_plantillas():
    for archivo, cabeceras in PLANTILLAS.items():
        ruta = os.path.join(DATA_DIR, archivo)
        rewrite = False
        if not os.path.exists(ruta):
            rewrite = True
        else:
            try:
                wb_exist = load_workbook(ruta)
                ws_exist = wb_exist.active
                existing_headers = [cell.value for cell in ws_exist[1]]
                if existing_headers != cabeceras:
                    rewrite = True
            except Exception:
                rewrite = True
        if rewrite:
            wb = Workbook()
            ws = wb.active
            ws.title = 'datos'
            ws.append(cabeceras)
            wb.save(ruta)
            print(f"Plantilla creada o actualizada: {archivo}")
        else:
            print(f"Plantilla válida: {archivo}")

if __name__ == '__main__':
    crear_plantillas()
